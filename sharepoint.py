from O365 import Account, FileSystemTokenBackend
from O365.excel import WorkBook

def modify_exel(file_instance):
    excel_file = WorkBook(file_instance)  # my_file_instance should be an instance of File.

    ws = excel_file.get_worksheet('my_worksheet')
    cella1 = ws.get_range('A1')
    cella1.values = 35
    cella1.update()

def read_exel(excel_file):
    #excel_file = WorkBook(file_instance)  # my_file_instance should be an instance of File.

    ws = excel_file.get_worksheet('CRITICAS')
    cella1 = ws.get_range('B13:K33')
    print(ws, cella1.values)
    # cella1 = ws.get_range('A1')
    # cella1.values = 35
    # cella1.update()


def load_workbook(account, filepath: str) -> WorkBook:
    parent, site, filepath = filepath.split(':')

    sp = account.sharepoint()

    # https://panaserviceperu.sharepoint.com/sites/BOT-PANASERVICE-REGISTRO

    main_dir = sp.get_site(parent, site)

    document_library = main_dir.get_document_library('.')
    file_obj = document_library.get_item_by_path(filepath)
    workbook = WorkBook(file_obj)
    return workbook
    

def read_by_id(drive, item_id):
    # item_id = 'D1ED06591598D29B!30531'  # Reemplaza con el ID real del archivo Excel

    # Descarga del archivo Excel para trabajar con él localmente
    excel_file = drive.get_item(item_id)
    excel_file.download('archivo_local.xlsx')

    # Aquí puedes agregar el código para modificar el archivo Excel usando una librería como 'openpyxl'
    # Por ejemplo, para agregar una nueva fila con contenido al final de una hoja existente:
    from openpyxl import load_workbook

    wb = load_workbook('archivo_local.xlsx')
    ws = wb.active  # Asume que trabajamos con la primera hoja

    # Agrega contenido a la nueva fila (reemplaza 'Nuevo Contenido' con el contenido real que deseas agregar)
    ws.append(['Nombre', 'Apellidos', 'Edad', 'Dirección', 'Tel', 'Email'])

    # Guarda los cambios en el archivo local
    wb.save('archivo_local.xlsx')

    # Sube el archivo modificado de vuelta a OneDrive
    excel_file.upload('archivo_local.xlsx')

    print("El archivo Excel ha sido actualizado y subido a OneDrive exitosamente.")


def ls_drive_folder(my_drive):
    # get some folders:
    root_folder = my_drive.get_root_folder()
    attachments_folder = my_drive.get_special_folder('attachments')

    # iterate over the first 25 items on the root folder
    for item in root_folder.get_items(limit=100):
        print(f"Nombre: {item.name}: ID: {item.object_id}")

        if item.is_folder:
            print(list(item.get_items(2)))  # print the first to element on this folder.

        elif item.is_file:
        
            if item.is_photo:
                print(item.camera_model)  # print some metadata of this photo
        
            elif item.is_image:
                print(item.dimensions)  # print the image dimensions
        
            else:
                # regular file:
                print(item.mime_type)


def main():
    # Configuración de las credenciales y el almacenamiento del token
    client_id = '683c3232-2013-4131-80fb-cb1285122d05'
    client_secret = None #'3d6198a0-6501-40fc-b40d-426128ffe95a'
    token_backend = FileSystemTokenBackend(token_path='.', token_filename='o365_token.json')
    # https://onedrive.live.com/edit?id=429D292F48EEBBBB!sacc1c1b40f4b4c6db069493a4f7d52cc&resid=429D292F48EEBBBB!sacc1c1b40f4b4c6db069493a4f7d52cc&cid=429d292f48eebbbb&ithint=file%2Cxlsx&redeem=aHR0cHM6Ly8xZHJ2Lm1zL3gvYy80MjlkMjkyZjQ4ZWViYmJiL0ViVEJ3YXhMRDIxTXNHbEpPazk5VXN3QjFMVDRtRUl0b044ZkNjcG5OU3YyelE_ZT1UUHR3Z24&migratedtospo=true&wdo=2
    # Autenticación con Microsoft Graph
    credentials = (client_id, client_secret)
    account = Account(credentials, token_backend=token_backend)

    # Si no estás autenticado, esto abrirá el proceso de autenticación en tu navegador predeterminado
    if not account.is_authenticated:
        account.authenticate(scopes=['basic', 'onedrive_all', 'sharepoint', 'sharepoint_dl'])

    # https://sicmaconsultores.sharepoint.com/sites/Ambiental2/Requerimientos de informacion V22 NDA1.xlsx
    # https://sicmaconsultores.sharepoint.com/sites/Ambiental/Requerimientos de informacion V22 NDA1.xlsx
    # https://sicmaconsultores.sharepoint.com/:x:/r/sites/Ambiental/_layouts/15/Doc.aspx?sourcedoc=%7B3E344534-B11E-4BA3-9557-8AD74B027390%7D&file=Requerimientos%20de%20informacion%20V22%20NDA1.xlsx&action=default&mobileredirect=true
    # Acceso a OneDrive y apertura del archivo Excel
    workbook = load_workbook(account, 'root:sites/Ambiental:/Requerimientos de informacion V22 NDA1.xlsx')
    #sharepoint = account.sharepoint()
    #drive = storage.get_default_drive()
    print(workbook)
    print(read_exel(workbook))

    #ls_drive_folder(drive)


if __name__ == '__main__':
    main()