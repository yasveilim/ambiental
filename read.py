# Importar el módulo openpyxl
import openpyxl

# Cargar el archivo excel
wb = openpyxl.load_workbook("Requerimientos de informacion V22 NDA1.xlsx")

# Obtener la hoja activa
sheet = wb.active

# Obtener el número de filas y columnas
max_row = 20 # sheet.max_row
max_col = 20 # sheet.max_column

# Recorrer todas las celdas y mostrar sus valores
for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        cell = sheet.cell(row=i, column=j)
        print(cell.value, end=" ")
    print()
